import openpyxl
import re
import time
import os
import pandas as pd


def ws_cell(ws, row_num, col_num):
    return ws.cell(row=row_num, column=col_num).value


def extract_data1(path):
    pattern = re.compile(r'^\d+[a-zA-Z]*\d*')
    wb_data = openpyxl.load_workbook(path)
    ws_input = wb_data.worksheets[0]
    ws_output = wb_data.create_sheet('提取数据', 1)

    def cell_length(row_num, col_num):
        return len(ws_cell(ws_input, row_num, col_num))

    for row_num in range(1, ws_input.max_row + 1):
        if re.match(pattern, ws_cell(ws_input, row_num, 1)) is not None:
            start_row = row_num
            break

    for row_num in range(start_row, ws_input.max_row + 1):
        if re.match(pattern, ws_cell(ws_input, row_num, 1)) is None:
            end_row = row_num
            break

    ls_row_num = []
    ls_row_name = []
    ls_row_cost = []
    ls_row_value = []
    ls_row_code = []
    ls_name_raw = []
    dict_code_name = {}
    for row_num in range(start_row, end_row):
        if cell_length(row_num, 1) >= cell_length(row_num + 1, 1):
            ls_row_num.append(row_num)
            ls_row_value.append(ws_cell(ws_input, row_num, 8))
            ls_row_cost.append(ws_cell(ws_input, row_num, 5))
            ls_row_code.append(ws_cell(ws_input, row_num, 1))
            ls_name_raw.append(ws_cell(ws_input, row_num, 2))
        else:
            dict_code_name[ws_cell(ws_input, row_num, 1)] = ws_cell(ws_input, row_num, 2)

    for index, code in enumerate(ls_row_code):
        ls_name = []
        for code_ in dict_code_name.keys():
            if re.match(code_, code) is not None:
                ls_name.append(dict_code_name[code_])
        ls_name.append(ls_name_raw[index])
        ls_row_name.append('-'.join(ls_name))

    ls_value_change = []
    for key_, value_ in dict_code_name.items():
        if value_.find('公允价值变动') + 1:
            print(key_)
            ls_value_change.append(key_)
        print(value_)
        print(value_.find('估值增值'))
        if value_.find('估值增值') + 1:
            print(key_)
            ls_value_change.append(key_)


    for index in reversed(range(len(ls_row_code))):
        code = ls_row_code[index]
        if ls_row_cost[index] is None or ls_row_cost[index] == '':
            ls_row_cost[index] = ls_row_value[index]
        if re.match('^2\d+', code) is not None:
            ls_row_num.pop(index)
            ls_row_cost.pop(index)
            ls_row_code.pop(index)
            ls_name_raw.pop(index)
            ls_row_value.pop(index)
            ls_row_name.pop(index)
        for code_change in ls_value_change:
            if re.match('^' + code_change + '\w+', code) is not None:
                ls_row_num.pop(index)
                ls_row_cost.pop(index)
                ls_row_code.pop(index)
                ls_name_raw.pop(index)
                ls_row_value.pop(index)
                ls_row_name.pop(index)

    data = pd.DataFrame({'行标': ls_row_num, '代码': ls_row_code, '名称': ls_row_name, '成本': ls_row_cost, '价值': ls_row_value})

    ws_output.cell(row=1, column=1).value = '代码'
    ws_output.cell(row=1, column=2).value = '名称'
    ws_output.cell(row=1, column=3).value = '成本'
    ws_output.cell(row=1, column=4).value = '市值'
    for index, item in enumerate(ls_row_code):
        row_output = index + 2
        ws_output.cell(row=row_output, column=1).value = item
        ws_output.cell(row=row_output, column=2).value = ls_row_name[index]
        if isinstance(ls_row_cost[index], str):
            ws_output.cell(row=row_output, column=3).value = float(ls_row_cost[index].replace(',', ''))
            ws_output.cell(row=row_output, column=4).value = float(ls_row_value[index].replace(',', ''))
        else:
            ws_output.cell(row=row_output, column=3).value = float(ls_row_cost[index])
            ws_output.cell(row=row_output, column=4).value = float(ls_row_value[index])
    wb_data.save(path[:-5] + '输出.xlsx')


time_start = time.time()
path_folder = 'D:\偿二代数据库\dist\\728估值表\权益估值表'
ls_path = os.listdir(path_folder)
ls_path = [path for path in ls_path if path[-7:] != '输出.xlsx']
print(ls_path)
for path_ in ls_path:
    extract_data1(path_folder + '/' + path_)
time_end = time.time()
print('用时{}秒'.format(time_end - time_start))
input()

