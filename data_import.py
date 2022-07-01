import openpyxl
from process_values import ws_cell
from process_values import max_row
from static_parameters import *
from Class_asset_data import AssetData
from process_values import cut_list


def complete_data(data: dict, col: list):
    data_complete = {}
    for key in col:
        if key in data.keys():
            data_complete[key] = data[key]
        else:
            data_complete[key] = None
    return data_complete


def bank_counter_party(bank_name_raw: str, dict_bank_name: dict):
    counter_party = str()
    if bank_name_raw.find('行')+1:
        counter_party = bank_name_raw[:bank_name_raw.find('行')+1]
    elif bank_name_raw.find('社保账户')+1:
        counter_party = bank_name_raw[:bank_name_raw.find('社保账户')+4]
    elif bank_name_raw.find('信用合作联社')+1:
        counter_party = bank_name_raw[:bank_name_raw.find('信用合作联社')+6]
    else:
        print("识别资产名失败")
    return dict_bank_name[counter_party]


def account_categorize(account_name: str):
    if account_name is not None and account_name.find('万能')+1:
        return '万能'
    else:
        return '传统'


def import_deposit(ws, dict_col_name_corr: dict, ls_colname: list):
    dict_data_raw = {}
    dict_data = {}
    ls_col_name_raw = [ws_cell(ws, 1, col_num) for col_num in range(1, ws.max_column+1)]
    for row_num in range(2, max_row(ws)+1):
        dict_row = {}
        ls_row = [ws_cell(ws, row_num, col_num) for col_num in range(1, ws.max_column+1)]
        dict_row_raw = dict(zip(ls_col_name_raw, ls_row))
        dict_data_raw[row_num-1] = dict_row_raw
        for col in ls_colname:
            if col in dict_col_name_corr:
                dict_row[col] = dict_row_raw[dict_col_name_corr[col]]
            elif col == '资产简称' or col == '资产全称':
                dict_row[col] = dict_row_raw['开户行'] + '-' + dict_row_raw['帐号'][-4:] + '-' + \
                                str(dict_row_raw['起息日'].year)[-2:] + \
                                str(dict_row_raw['起息日'].month).rjust(2, '0') + \
                                str(dict_row_raw['起息日'].day).rjust(2, '0')
            elif col == '资产大类':
                dict_row[col] = '银行存款'
            elif col == '账户':
                dict_row[col] = account_categorize(dict_row_raw['账户'])
            elif col == '交易对手':
                dict_row[col] = bank_counter_party(dict_row_raw['开户行'], dict_bank_counter_party)
            elif col == '银行资本充足率':
                dict_row[col] = dict_row_raw['资本充足率（%）']/100
            elif col == '资产五大类分类':
                dict_row[col] = '固定收益类资产'
            else:
                dict_row[col] = None
        # print(dict_row)
        dict_data[row_num-1] = complete_data(dict_row, ls_col_all)
    return dict_data


def import_current_deposit(ws, dict_col_name_corr: dict, dict_col_value: dict, ls_colname: list):
    dict_data_raw = {}
    dict_data = {}
    ls_col_name_raw = [ws_cell(ws, 3, col_num) for col_num in range(1, ws.max_column+1)]
    for row_num in range(4, max_row(ws)+1):
        dict_row = {}
        ls_row = [ws_cell(ws, row_num, col_num) for col_num in range(1, ws.max_column+1)]
        dict_row_raw = dict(zip(ls_col_name_raw, ls_row))
        dict_data_raw[row_num-3] = dict_row_raw
        for col in ls_colname:
            if col in dict_col_name_corr:
                dict_row[col] = dict_row_raw[dict_col_name_corr[col]]
            elif col in dict_col_value:
                dict_row[col] = dict_col_value[col]
            elif col == '账户':
                dict_row[col] = account_categorize(dict_row_raw['机构名称'])
            elif col == '交易对手':
                dict_row[col] = bank_counter_party(dict_row_raw['帐户信息'], dict_bank_counter_party)
            else:
                dict_row[col] = None
        # print(dict_row)
        dict_data[row_num-1] = complete_data(dict_row, ls_col_all)
    return dict_data


def import_invest_asset(ws, ls_col_name):
    ls_col_name_raw = [ws_cell(ws, 1, col_num) for col_num in range(1, ws.max_column+1)]
    dict_name_col = {col_name: ls_col_name_raw.index(col_name)+1 for col_name in ls_col_name}
    # row_num = 2
    # print({col_name: ws_cell(ws, row_num, dict_name_col[col_name]) for col_name in ls_col_name})
    dict_data = {row_num-1: {col_name: ws_cell(ws, row_num, dict_name_col[col_name]) for col_name in ls_col_name} for
                 row_num in range(2, max_row(ws)+1)}
    return dict_data


def import_non_invest_data(ws, ls_col_name):
    ls_col_name_raw = [ws_cell(ws, 2, col_num) for col_num in range(1, ws.max_column+1)]
    dict_name_col = {col_name: ls_col_name_raw.index(col_name)+1 for col_name in ls_col_name}
    # row_num = 2
    # print({col_name: ws_cell(ws, row_num, dict_name_col[col_name]) for col_name in ls_col_name})
    dict_data = {row_num-2: {col_name: ws_cell(ws, row_num, dict_name_col[col_name]) for col_name in ls_col_name} for
                 row_num in range(3, max_row(ws)+1)}
    for i in dict_data.keys():
        dict_data[i]['资产全称'] = dict_data[i]['资产简称']
        dict_data[i]['购买成本'] = dict_data[i]['认可价值']
        if dict_data[i]['资产大类'] == '投资性房地产':
            dict_data[i]['资产五大类分类'] = '不动产类资产'
        else:
            dict_data[i]['资产五大类分类'] = '非投资资产'
        if dict_data[i]['交易对手'] is None:
            dict_data[i]['交易对手'] = '无交易对手'

    return dict_data


wb_deposit_data = openpyxl.load_workbook('资产端数据/定期存款-2022年5月.xlsx', data_only=True)
wb_current_deposit_date = openpyxl.load_workbook('资产端数据/活期存款科目余额表.xlsx')
wb_invest_asset = openpyxl.load_workbook('资产端数据/530汇总表.xlsx')
wb_non_invest_asset = openpyxl.load_workbook('资产端数据/数据模板与说明（非投资资产）.xlsx')

ws_deposit_data = wb_deposit_data['定期存款和存出资本保证金应收利息明细']
ws_current_deposit_data = wb_current_deposit_date.worksheets[0]
ws_invest_asset = wb_invest_asset.worksheets[0]
ws_non_invest_asset = wb_non_invest_asset

data_deposit = import_deposit(ws_deposit_data, dict_deposit_col_name_self, ls_deposit_col_name)
data_current_deposit = import_current_deposit(ws_current_deposit_data, dict_current_deposit_col_name_self,
                                              dict_current_deposit_col_value, ls_current_deposit)
data_invest_asset = import_invest_asset(ws_invest_asset, ls_col_all)
data_non_invest_asset = import_non_invest_data(wb_non_invest_asset['填写模板'], ls_col_non_invest)
ls_data_deposit = [complete_data(i, ls_col_all) for i in data_deposit.values()]
ls_data_current_deposit = [complete_data(i, ls_col_all) for i in data_current_deposit.values()]
ls_data_invest = [complete_data(i, ls_col_all) for i in data_invest_asset.values()]
ls_data_non_invest = [complete_data(i, ls_col_all) for i in data_non_invest_asset.values()]
ls_data_all = [*ls_data_invest, *ls_data_deposit, *ls_data_current_deposit, *ls_data_non_invest]
ls_pure_data_all = [list(item.values()) for item in ls_data_all]

ls_sd_data = []
ls_labeled_data = []
ls_dict_sd_data = []
i = 1
for data in ls_data_all:
    asset = AssetData(data)
    ls_sd_data.append(list(asset.sd_data.values()))
    ls_dict_sd_data.append(asset.sd_data)
    ls_labeled_data.append(list(asset.labeled_data.values()))
    i += 1
ls_col_sd_data = list(asset.sd_data.keys())
ls_col_labeled_data = list(asset.labeled_data.keys())

'''###########################################################################################################'''
# complete_data(data_deposit[1], ls_col_all)
#
# asset = AssetData(data_invest_asset[831])
# # ls_result = asset.data_check()
# from database import MySqlConnection
# sb_solv = MySqlConnection(host="localhost", user="root", password="19981027phy", database="solvency2")
# # sb_solv.connect_db()
# sb_solv.insert("data_raw", asset.sd_data.keys(), list(asset.sd_data.values()))
