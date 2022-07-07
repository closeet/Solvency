import openpyxl
import pickle
from process_values import ws_cell
from process_values import max_row


def serialize(ls, filename):
    with open(filename, 'wb') as file:
        pickle.dump(ls, file)


def deserialize(filename):
    with open(filename, 'rb') as file:
        return pickle.load(file)


wb_data_rule = openpyxl.load_workbook('数据规则.xlsx')
ws_rule_bank = wb_data_rule['存款交易对手']
ws_rule_null = wb_data_rule['字段空与非空']
ws_rule_type = wb_data_rule['数据类型']
ws_rule_valid = wb_data_rule['字段取值']
ws_rule_k = wb_data_rule['k1,k2']
ws_rule_others = wb_data_rule['其他']
ws_penetration_exempt = wb_data_rule['豁免穿透资产类型']
ws_vague_value = wb_data_rule['模糊字段']

dict_deposit_col_name_self = {'购买成本': '存款金额', '认可价值': '认可价值', '应收利息': '应收利息', '存款银行类型': '银行类型', '资产类型': '存款类型'}
ls_deposit_col_name = ['资产简称', '资产全称', '资产大类', '资产类型', '交易对手', '购买成本', '认可价值', '应收利息', '账户', '存款银行类型', '银行资本充足率',
                       '资产五大类分类']
dict_current_deposit_col_name_self = {'资产简称': '帐户信息', '资产全称': '帐户信息', '购买成本': '余额', '认可价值': '余额'}
dict_current_deposit_col_value = {'资产大类': '现金及流动性管理工具', '资产类型': '活期存款', '资产五大类分类': '流动性资产'}

ls_current_deposit = ['资产简称', '资产全称', '资产大类', '资产类型', '交易对手', '购买成本', '认可价值', '账户', '资产五大类分类']

dict_bank_counter_party = {ws_cell(ws_rule_bank, row_num, 1): ws_cell(ws_rule_bank, row_num, 2) for row_num in
                           range(1, ws_rule_bank.max_row + 1) if ws_cell(ws_rule_bank, row_num, 1) is not None}

ls_col_all = [ws_cell(ws_rule_type, row_num, 2) for row_num in range(1, max_row(ws_rule_type) + 1)]
ls_col_non_invest = ['资产简称', '资产大类', '资产类型', '交易对手', '认可价值', '应收利息', '账户', '所在城市', '投资时间', '计量属性',
                     '账面价值', '是否是享受各级政府保费补贴的业务', '账龄']

ls_core_city = [ws_cell(ws_rule_others, row_num, 3) for row_num in range(2, ws_rule_others.max_row + 1) if
                ws_cell(ws_rule_others, row_num, 3) is not None]
dict_country_currency = {ws_cell(ws_rule_others, row_num, 6): ws_cell(ws_rule_others, row_num, 7) for row_num in
                         range(2, ws_rule_others.max_row + 1) if ws_cell(ws_rule_others, row_num, 6) is not None}
dict_currency_country = {currency: [item[0] for item in dict_country_currency.items() if item[1] == currency]
                         for currency in set(dict_country_currency.values())}
dict_country_development = {ws_cell(ws_rule_others, row_num, 6): ws_cell(ws_rule_others, row_num, 8) for row_num in
                            range(2, ws_rule_others.max_row + 1) if ws_cell(ws_rule_others, row_num, 6) is not None}
ls_country_developing = [items[0] for items in dict_country_development.items() if items[1] == '新兴市场']
dict_risk_type = {ws_cell(ws_rule_others, row_num, 4): ws_cell(ws_rule_others, row_num, 5) for row_num in
                  range(2, ws_rule_others.max_row + 1) if ws_cell(ws_rule_others, row_num, 4) is not None}
dict_type_general = {ws_cell(ws_rule_others, row_num, 9): ws_cell(ws_rule_others, row_num, 10) for row_num in
                     range(2, ws_rule_others.max_row + 1) if ws_cell(ws_rule_others, row_num, 9) is not None}
dict_rule_null = {ws_cell(ws_rule_null, row_num, 1): [ws_cell(ws_rule_null, 1, col) for col in
                                                      range(2, ws_rule_null.max_column + 1)
                                                      if ws_cell(ws_rule_null, 1, col) is not None and
                                                      ws_cell(ws_rule_null, row_num, col) == 1]
                  for row_num in range(2, ws_rule_null.max_row + 1) if ws_cell(ws_rule_null, row_num, 1) is not None}
dict_rule_type = {ws_cell(ws_rule_type, row_num, 2): ws_cell(ws_rule_type, row_num, 3) for row_num in
                  range(1, ws_rule_type.max_row + 1) if ws_cell(ws_rule_type, row_num, 2) is not None}
dict_rule_valid = {ws_cell(ws_rule_valid, 1, col_num): [ws_cell(ws_rule_valid, row_num, col_num) for row_num in
                                                        range(1, ws_rule_valid.max_row + 1) if
                                                        ws_cell(ws_rule_valid, row_num, col_num) is not None]
                   for col_num in range(1, ws_rule_valid.max_column + 1) if
                   ws_cell(ws_rule_valid, 1, col_num) is not None}
dict_rule_lower = {ws_cell(ws_rule_type, row_num, 2): ws_cell(ws_rule_type, row_num, 4) for row_num in
                   range(1, ws_rule_type.max_row + 1) if ws_cell(ws_rule_type, row_num, 4) is not None}
dict_rule_upper = {ws_cell(ws_rule_type, row_num, 2): ws_cell(ws_rule_type, row_num, 5) for row_num in
                   range(1, ws_rule_type.max_row + 1) if ws_cell(ws_rule_type, row_num, 5) is not None}
dict_rule_compare = {'发行银行核心一级资本充足率': '发行银行一级资本充足率', '发行银行一级资本充足率': '发行银行资本充足率', '发行保险公司核心偿付能力充足率': '发行保险公司综合偿付能力充足率'}
dict_rule = {'not_null': dict_rule_null, 'type': dict_rule_type, 'valid': dict_rule_valid, 'lower': dict_rule_lower,
             'upper': dict_rule_upper, 'compare': dict_rule_compare}
ls_null_values = {'无', '-'}

dict_asset_type_proportion = {'权益': 0.25, '房地产': 0.25, '其他': 0.25, '境外': 0.15}
ls_penetration_exempt_type = [ws_cell(ws_penetration_exempt, row_num, 1) for row_num in
                              range(2, ws_penetration_exempt.max_row + 1) if ws_cell(ws_penetration_exempt, row_num, 1)
                              is not None]

dict_vague_value_bank = {ws_cell(ws_vague_value, 2, col_num):
                             [ws_cell(ws_vague_value, row_num, col_num) for row_num in
                              range(2, ws_vague_value.max_row + 1)
                              if ws_cell(ws_vague_value, row_num, col_num) is not None]
                         for col_num in range(1, ws_vague_value.max_column+1) if ws_cell(ws_vague_value, 1, col_num) == '银行类型'}
dict_vague_asset_5_types = {ws_cell(ws_vague_value, 2, col_num):
                             [ws_cell(ws_vague_value, row_num, col_num) for row_num in
                              range(2, ws_vague_value.max_row + 1)
                              if ws_cell(ws_vague_value, row_num, col_num) is not None]
                         for col_num in range(1, ws_vague_value.max_column+1) if ws_cell(ws_vague_value, 1, col_num) == '五大类'}

dict_vague_value = {'发行银行类型': dict_vague_value_bank, '存款银行类型': dict_vague_value_bank, '资产五大类分类': dict_vague_asset_5_types}
